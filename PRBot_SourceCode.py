#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jul  1 14:04:55 2019

@author: SplitShiftKing
"""

import tkinter as tk
from tkinter import filedialog as fd, Entry, messagebox as mb, ttk
import pandas as pd
from openpyxl import load_workbook
import os
import time
from datetime import timedelta
from openpyxl.styles import PatternFill
import requests as rq
from bs4 import BeautifulSoup
import unicodedata

# Change working directory
path = ("/Users/SplitShiftKing/Documents/Career"
        "/Projects/ProjectRunway/PR_Bot/WithClasses")
os.chdir(path)

# Create root window
InWin = tk.Tk()


# Create frame class
class Page(tk.Frame):
    """class for creating page template that has frame properties"""
    
    # Dictionary for housing the different pages of the app (frames)
    pages = {}
    # Page/frame key
    page_count = 0
    # Instantiated to determine if 'back' is being pressed from the final page where the tag search occurs.
    from_fin = 0

    def __init__(self, master):
        tk.Frame.__init__(self, master)
        # Q: self=Frame(master), but how does tk.Frame know how to use master?
        # Q: self.page_count += 1 references the class variable, but can't change it. Why?

        # Update class variable
        Page.page_count += 1
        Page.pages[Page.page_count] = self
        Page.pages[Page.page_count].grid(row=1, column=1, sticky='NSEW')

    # Add button, label, or entry widget to a page
    def add_button(self, text, function, x=0, y=0, sticky=None, L=0, R=0, U=0, D=0, formatting=None):
        button = tk.Button(self, text=text, command=function, font=formatting)
        button.grid(row=x, column=y, sticky=sticky, padx=(L, R), pady=(U, D))
        return button

    def add_label(self, text, x=0, y=0, sticky=None, L=0, R=0, U=0, D=0, formatting=None):
        label = tk.Label(self, text=text, font=formatting)
        label.grid(row=x, column=y, sticky=sticky, padx=(L, R), pady=(U, D))
        return label

    def entry_widget(self, x=0, y=0, sticky=None, L=0, R=0, U=0, D=0):
        entry = Entry(self)
        entry.grid(row=x, column=y, sticky=sticky, padx=(L, R), pady=(U, D))
        return entry

    def refresh(self):
        """Reset current page to its default state"""

        # Destroy current page
        self.destroy()
        # Delete stored tags for current brand
        TagPage.saved_tags[HomePage.bran_list_0[TagPage.brand]] = {}
        # Reset widget placements
        TagPage.r = 0
        TagPage.c = 0
        TagPage.t = 0
        TagPage.l = 0
        TagPage.k = 1
        TagPage.pwl = 0
        # change page_count before it's incremented in initialization to avoid getting a duplicate page.
        Page.page_count -= 1
        # Recreate page
        TagPage(InWin)

        ## print('\nrefreshed:')
        ## print('brand = ', TagPage.brand)
        ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
        ## print('page count = ', Page.page_count)
        ## print('pages = ', Page.pages)
        ## print('tags = {}'.format(TagPage.saved_tags))

    def back(self):
        """Return to previous page"""

        Page.page_count = Page.page_count-1
        Page.pages[Page.page_count].tkraise()
        # If back pressed to go to HomePage (when page_count=2), don't decrement brand so that brand doesn't go to -1, in which case bran_list_0[brand] will equal 4thStreetWine.
        if Page.page_count == 1:
            pass
        # If back pressed from final page, decrement from_fin class variable, but don't decrement brand because we are still on the same brand, unlike when next is pressed.
        elif Page.from_fin == 1:
            Page.from_fin -= 1
            # Reset bran_list_1 which was altered when finish was pressed.
                # Q: I don't know why bran_list_0 kept getting altered when it was a list, but making it a tuple solved the problem.
            HomePage.bran_list_1 = list(HomePage.bran_list_0)

            ## print('\nafter pressing back from SearchPage:') 
            ## print('add unlisted brands:', bool(SearchPage.v.get()))
            ## print('v value = ', SearchPage.v.get())
            ## print('from_fin value = ', Page.from_fin)
            ## print('bran_list_1 restored = ', HomePage.bran_list_1)
            ## print('bran_list_0 = ', HomePage.bran_list_0)
            ## print('\nbrand = ', TagPage.brand)
            ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
            ## print('page count = ', Page.page_count)
            ## print('pages = ', Page.pages)
        else:
            # Decrements brand to ensure that previous brand's dictionary is being altered
            TagPage.brand -= 1

            ## print('\nbrand = ', TagPage.brand)
            ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
            ## print('page count = ', Page.page_count)
            ## print('pages = ', Page.pages)
            ## print('tags = {}'.format(TagPage.saved_tags))



class HomePage(Page):
    """class for creating home page"""

    # Variable for storing list of brands being looked for
    bran_list_0 = None
    # Variable for storing altered list of brands
    bran_list_1 = None
    # Dictionary for storing tags associated with each brand
    tags = {}
    check_blanks = {}

    def __init__(self, master):
        # Run Page class initialization method for standard page template
        Page.__init__(self, master)
        Page.add_button(self, 'Browser', self.open_file, 0, 0)
        Page.add_button(self, 'Enter Tags', self.enter_tags, 1, 0)

    def open_file(self):
        """Open excel file, filter it, then output edited version"""

        # Opens file browser and stores file that you select
        ExcelFile = fd.askopenfile()
        # Returns file path
        path = r'' + ExcelFile.name
        # Turn excel file into a data frame
        df = pd.read_excel(path)
        # Filters frame so only working with 'online news coverage' category
        is_online = df['Category Name'] == 'Online News Coverage: Project Runway SA'
        HomePage.dfR_filt = df[(is_online)]
        # Return list of brands being looked for. Had to make it a tuple so it doesn't get altered later when I iterate over it to alter bran_list_1.
        HomePage.bran_list_0 = tuple(HomePage.dfR_filt['Sub-Category Name'].unique())
        # Create nested dictionary that tag widgets will go into
        for brand_1st in HomePage.bran_list_0:
            HomePage.tags[brand_1st] = {}
        # Create nested dictionary that tag widget values will be stored in
        for brand_2nd in HomePage.bran_list_0:
            HomePage.check_blanks[brand_2nd] = {}
        # Copy of variable bran_list_0 that will be altered depending on which brands are selceted
        HomePage.bran_list_1 = list(HomePage.dfR_filt['Sub-Category Name'].unique())
        # Save frame as excel file with new name, then load that excel file using openpyxl
        new_path = path.replace('.xlsx', '_online') + '.xlsx'
        HomePage.dfR_filt.to_excel(new_path, engine='openpyxl')
        HomePage.BrandWB = load_workbook(new_path)
        HomePage.ws1 = HomePage.BrandWB['Sheet1']

    def enter_tags(self):
        """Creates new page that moves us to next section of application where brand tags are entered"""

        # First time running through program, if-statement will raise error because saved_tags won't exist yet. But if HomePage navigated to via back, then saved_tags will already exist. 
        try:
            if bool(TagPage.saved_tags[HomePage.bran_list_0[TagPage.brand]]):
                ## print('\nRaised Page:')
                Page.page_count += 1
                Page.pages[Page.page_count].tkraise()

                ## print('brand = ', TagPage.brand)
                ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
                ## print('page count = ', Page.page_count)
                ## print('pages = ', Page.pages)
                ## print('tags = {}'.format(TagPage.saved_tags))
            else:
                ## print('\ntry bloc, New Page:')
                TagPage.r = 0
                TagPage.c = 0
                TagPage.t = 0
                TagPage.l = 0
                TagPage.k = 1
                TagPage.pwl = 0
                TagPage(InWin)

                ## print('brand = ', TagPage.brand)
                ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
                ## print('page count = ', Page.page_count)
                ## print('pages = ', Page.pages)
                ## print('tags = {}'.format(TagPage.saved_tags))

        except:
            ## print('\nexcept bloc, New Page:')
            TagPage.r = 0
            TagPage.c = 0
            TagPage.t = 0
            TagPage.l = 0
            TagPage.k = 1
            TagPage.pwl = 0
            TagPage(InWin)

            ## print('brand = ', TagPage.brand)
            ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
            ## print('page count = ', Page.page_count)
            ## print('pages = ', Page.pages)


class TagPage(Page):
    """Creates pages where tags are entered"""

    # Dictionary where tags are saved when 'save' button pressed
    saved_tags = {}
    # What brand is being worked on
    brand = 0
    # Space in between widgets
    step = 4
    # Previous widget length. Used to place widget right distance from west 0,0 position
    pwl = 0
    # System for storing entry_widget values according to grid rows and columns
    #  and_function row value
    r = 0
    #  and_function column value
    c = 0
    #  or_function row value
    t=0
    #  or_function increment
    l=0
    k=1


    def __init__(self, master):    
        Page.__init__(self, master)

        brandLabel = Page.add_label(self,
                                   'For {}, search for:'.format(HomePage.bran_list_0[TagPage.brand]),
                                   x=0, y=0, sticky='W')

        andButton = Page.add_button(self, text='And', function=self.and_function, sticky='W',
                                   L=TagPage.spacing(self, brandLabel),
                                   formatting='Arial 13 bold')

        orButton = Page.add_button(self, text='Or', function=self.or_function, sticky='W',
                                  L=TagPage.spacing(self, andButton) + TagPage.step,
                                  formatting='Arial 13 bold')

        save_tagsButton = Page.add_button(self, text='Save', function=self.save_tags, sticky='W',
                                        L=TagPage.spacing(self, orButton) + 2*(TagPage.step),
                                        formatting='Arial 13 bold')

        nextButton = Page.add_button(self, text='Next', function=self.next_brand, sticky='W',
                                        L=TagPage.spacing(self, save_tagsButton) + 3*(TagPage.step),
                                        formatting='Arial 13 bold')

        finishButton = Page.add_button(self, text='Finish', function=self.finish, sticky='W',
                                        L=TagPage.spacing(self, nextButton) + 4*(TagPage.step),
                                        formatting='Arial 13 bold')

        refreshButton = Page.add_button(self, text='Refresh', function=self.refresh, sticky='W',
                                        L=TagPage.spacing(self, finishButton) + 5*(TagPage.step),
                                        formatting='Arial 13 bold')

        Page.add_button(self, text='Back', function=self.back, sticky='W',
                        L=TagPage.spacing(self, refreshButton) + 6*(TagPage.step),
                        formatting='Arial 13 bold')

        # Starting entry widget
        TagPage.entry_widget(self, 1, 0, sticky='W')


    def entry_widget(self, x, y, key_r=0, key_c=0, sticky=None, L=0, R=0, U=0, D=0):
        """Place widgets associated with brand x and their inputs into the tags dictionary,
        where the key is brand x followed by the position of that entry widget in the grid of the page,
        for example, {Project Runway:{00: Entry!1, 01: Entry!2}}.
        """
        HomePage.tags[HomePage.bran_list_0[TagPage.brand]][str(key_r)+str(key_c)] = tk.Entry(self)
        HomePage.tags[HomePage.bran_list_0[TagPage.brand]][str(key_r)+str(key_c)].grid(row=x, column=y, sticky=sticky, padx=(L, R), pady=(U, D))

    def spacing(self, widget):
        """return length of 'widget' in pixels from x=0"""
        
        # Updates widget variables that change
        self.update_idletasks()
        TagPage.pwl = TagPage.pwl + widget.winfo_width()
        return TagPage.pwl #not entirely necessary

    def and_function(self):
        """Add entry widgets horizontally when 'and' is pressed"""
        
        TagPage.c = TagPage.c+1
        # TagPage.c==1: If and is pressed for the first time
        if TagPage.c==1:

            # How fixed widget length was calculated:
                # wl = tags[bran_list_0[brand]]['00'].winfo_width()
                # print(wl) = 192

            Page.add_label(self, text='and', x=TagPage.r+1, y=0, sticky='W', L=192)

            # How label length was calculated:
                # LL = label.winfo_width()
                # print(LL) = 34
                # L = wl + LL = 226

            TagPage.entry_widget(self, x=TagPage.r+1, y=0, key_r=TagPage.r, key_c=TagPage.c,
                                sticky='W', L=226)
        else:

            Page.add_label(self, text='and', x=TagPage.r+1,y=0, sticky='W', L=(226*TagPage.c)-34)
            TagPage.entry_widget(self, x=TagPage.r+1, y=0, key_r=TagPage.r, key_c=TagPage.c,
                                sticky='W', L=226*TagPage.c)
    def or_function(self):
        """Add entry widgets vertically when 'or' is pressed"""

        TagPage.r = TagPage.r + 2
        TagPage.t = TagPage.t + 1
        TagPage.l = TagPage.l + 1
        TagPage.k = TagPage.k + 1
        # Reset and widget row value
        TagPage.c = 0

        #TagPage.t==1: If or is pressed for the first time
        if TagPage.t==1:
            Page.add_label(self, text ='or', x=TagPage.t+1,y=0, sticky='W', L=(192/2)-10)
            TagPage.entry_widget(self, x=TagPage.t+2,y=0, key_r=TagPage.r, key_c=TagPage.c,
                                sticky='W')
        else:
            Page.add_label(self, text ='or', x=TagPage.t+TagPage.l, y=0, sticky='W', L=(192/2)-10)
            TagPage.entry_widget(self, x=TagPage.t+TagPage.k,y=0, key_r=TagPage.r, key_c=TagPage.c,
                                sticky='W')


    def save_tags(self):
        """Place tags associated with brand x into the saved_tags dictionary,
        where the key is brand x followed by the position of that entry widget in the grid of the page,
        for example, {Project Runway:{00: Project Runway, 01: South Africa}.
        """

        # Replacing widgets with the values contained in them
        for position_2 in HomePage.tags[HomePage.bran_list_0[TagPage.brand]].keys():
                HomePage.check_blanks[HomePage.bran_list_0[TagPage.brand]][position_2] = HomePage.tags[HomePage.bran_list_0[TagPage.brand]][position_2].get()
        # True if tags are empty
            # NB: to add condition that counts list of keys and checks if > 2. So that message doesn't appear if first widget left blank.
        if '' in list(HomePage.check_blanks[HomePage.bran_list_0[TagPage.brand]].values()):
            mb.showerror('Error', 'There can be no blank spaces!')
        else:
            TagPage.saved_tags = HomePage.check_blanks

        ## print('\nsaved:')
        ## print('brand = ', TagPage.brand)
        ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
        ## print('page count = ', Page.page_count)
        ## print('pages = ', Page.pages)
        ## print('tags = {}'.format(TagPage.saved_tags))

    def next_brand(self):
        """Move onto next brand in HomePage.bran_list_0"""

        # Note to self: Use bool to evaluate true if something is in dictionary. Use not to evaluate true if nothing is in dictionary.
        # If no tags saved for current brand, do nothing. Otherwise, update current brand (go onto next brand).
        if not TagPage.saved_tags[HomePage.bran_list_0[TagPage.brand]]:
            pass
        # When last element of bran_list_0 has been reached, do nothing.
        elif TagPage.brand == len(startPage.pages[1].bran_list_0) - 1:
            pass
        else:
            # Go to next brand
            TagPage.brand += 1
            # Reset TagPage widget positions and variables
            TagPage.r = 0
            TagPage.c = 0
            TagPage.t = 0
            TagPage.l = 0
            TagPage.k = 1
            TagPage.pwl = 0

            # If next brand already has tags saved (meaning its page has already been created), then simply raise that page. Otherwise create new TagPage page.
            if bool(TagPage.saved_tags[HomePage.bran_list_0[TagPage.brand]]):
                ## print('\nRaised Page:')
                Page.page_count += 1
                Page.pages[Page.page_count].tkraise()

                ## print('brand = ', TagPage.brand)
                ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
                ## print('page count = ', Page.page_count)
                ## print('pages = ', Page.pages)
                ## print('tags = {}'.format(TagPage.saved_tags))
            else:
                ## print('\nNew Page:')
                TagPage(InWin)

                ## print('brand = ', TagPage.brand)
                ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
                ## print('page count = ', Page.page_count)
                ## print('pages = ', Page.pages)
                ## print('tags = {}'.format(TagPage.saved_tags))

    def finish(self):
        """Attached to 'finish' button:
            create new page that takes us to final section of interface where search occurs
        """
        #Error 3 Solution: recognize when there is a brand page following the current brand page and in that instance, don't overwrite it with SearchPage. Or rather probably easier to keep SearchPage when created, and only raise it thereafter, adding other tag pages in between. the same way that HomePage stays open but just hidden.
            #Let's start by keeping it and adding other pages round it

        SearchPage(InWin)

        ## print('bran_list_1 altered = ', HomePage.bran_list_1)
        ## print('bran_list_0 = ', HomePage.bran_list_0)
        ## print('\nbrand = ', TagPage.brand)
        ## print('current brand name = ', HomePage.bran_list_0[TagPage.brand])
        ## print('page count = ', Page.page_count)
        ## print('pages = ', Page.pages)

class SearchPage(Page):
    """Class for final page of interface"""

    # Variable that will be attached to radio button to determine if unlisted brands should be searched for and added to spreadsheet.a
    v = tk.IntVar()
    # Determines what range of cells will be worked on in the spreadsheet
    searchRange = None
    # Color definitions
        # Indicates all tags present
    green_fill = PatternFill(start_color='00008000',
                             end_color='00008000',
                             fill_type='solid')
        # Indicates not all tags present
    red_fill = PatternFill(start_color='00FF0000',
                           end_color='00FF0000',
                           fill_type='solid')
        # Indicates case where exception occured due to network problem or anything else
    blue_fill = PatternFill(start_color='000000FF',
                           end_color='000000FF',
                           fill_type='solid')

    def __init__(self, master):
        if not TagPage.saved_tags[HomePage.bran_list_0[TagPage.brand]]:
            mb.showerror('Error', 'You did not save tags!')
        else:
            Page.__init__(self, master)
            # Reset pwl so that its value from tag pages doesn't carry over
            TagPage.pwl = 0 #Check to make sure it doesn't affect anything else that requires pwl like widget placements on TagPages
            SearchPage.v.set(None)
            # Changes from_fin class variable in Page class to indicate that we are now on the final page.
            Page.from_fin = 1
            ## print('\nafter pressing finish:')
            ## print('from_fin value = ', Page.from_fin)

            Page.add_label(self, x=1, text='Entries: {}'.format(HomePage.ws1.max_row-1),
                                         sticky='W', L=4)
            # Range label + range entry_widget
            Page.add_label(self, text = 'Search Range (e.g., 2:11): ', x=2,
                          sticky='W', L=4)
            SearchPage.searchRange = Page.entry_widget(self, x=2, y=0, sticky='E', L=180, R=4)

            # Search button
            Page.add_button(self, text='Search', formatting='Arial 13 bold',
                           function=self.search, x=4, sticky='W', L=5, R=4)

            # Back button
            Page.add_button(self, text='Back', formatting='Arial 13 bold',
                          function=self.back, x=1, y=1, sticky='E', L=150, R=4)

            # Progress bar for showing search completion percentage
            SearchPage.progress_bar = ttk.Progressbar(self, orient='horizontal',
                                           mode='determinate', length=200)
            SearchPage.progress_bar.grid(row=5, column=0, sticky='W', pady=(5,0))
            SearchPage.progress_label = tk.Label(self, text='')
            SearchPage.progress_label.grid(row=5, column=0, padx=(100,0))

            # Radio buttons: change value of v to 1 (true) or 0 (false) if yes or no are pressed respectively
            unlistedBrandButton = Page.add_label(self, text='Add Unlisted Brands:', x=3, sticky='W', L=4)
            actionButton1 = tk.Radiobutton(self, variable=SearchPage.v, text='yes', value=1) #Think that makes default variable true
            actionButton1.grid(row=3, sticky='W', padx=(TagPage.spacing(self, unlistedBrandButton),0))
            actionButton2 = tk.Radiobutton(self, variable=SearchPage.v, text='no', value=0)
            actionButton2.grid(row=3, sticky='W', padx=(TagPage.spacing(self, actionButton1),0))
            # Displays how long search took
            durationText = tk.Label(self, text='Duration:')
            durationText.grid(row=5, column=1, sticky='W')
            SearchPage.durationLabel = tk.Label(self, text='')
            SearchPage.durationLabel.grid(row=5, column=1, sticky='E', padx=(0,0))
            # Remove brands that have no tags from bran_list_1 so that they aren't included in the search
            for brand_7th in HomePage.bran_list_0:
                if (not TagPage.saved_tags[brand_7th]):
                    #print('if: ', brand_7th)
                    HomePage.bran_list_1.remove(brand_7th)
                else:
                    pass

    def search(self):
        """Go row by row, searching articles in the article column for
        the tags associated with each brand in the brand column
        """

        # Store time when search function starts
        start_time = time.monotonic()
        # Conditions dictionary: stores true or false variable for each brand tag depending on if it was found on the website article in quesiton
        cons_dict = {}
        for brand_8th in HomePage.bran_list_0:
            cons_dict[brand_8th] = {}
        # Sum of conditions dictionary: stores true or false variable for each brand depending on the sum of its individual tag boolean values. E.g., true and false or true = true. true and true and false = false.
        sum_cons = {}

        ## print('\nafter pressing search:\n')
        ## print('conditions dictionary = ', cons_dict)
        ## print('sum of conditions dictionary = ', sum_cons)

        # Find column value of 'Article URL' and 'Sub-category Name' headings in spreadsheet in order to reference them
        for row_cells in HomePage.ws1.iter_rows():
            for cell in row_cells:
                if cell.value=='Article URL':
                    art_col=cell.column
                    break
                else:
                    pass
            break

        for row_cells in HomePage.ws1.iter_rows():
            for cell in row_cells:
                if cell.value=='Sub-Category Name':
                    brand_col=cell.column
                    break
                else:
                    pass
            break
        ## print('article column = ', art_col)
        ## print('brand column = ', brand_col)

        def cs(x):
            """Changes string by removing single spaces and lowering case"""

            return x.replace(' ','').lower()

        # Take entry for 'Search Range' and use as loop range
        loop_range = SearchPage.searchRange.get().split(':') #check to make sure that entry format okay
        ## print('loop range = ', loop_range)
        loop_range = [int(q) for q in loop_range]

        # Variables for setting progress bar value. loop_range[1] - loop_range[0] = number of articles being scanned.
        delta = (1/(loop_range[1]-loop_range[0]))*100 # Percentage value each article incremenets the bar by
        int_var = tk.IntVar() #What does this do again? *
        
        # Search function algorithm. Heart of program.
        for num, cell in enumerate(range(loop_range[0], loop_range[1])):
            try:
                store_site = rq.get(HomePage.ws1[art_col+str(cell)].value, headers={'User-Agent': 'Mozilla/5.0'})
                # Raises exception for 4xx and 5xx errors which aren't caught normally  
                store_site.raise_for_status()
                ## print('\nsite = ', HomePage.ws1[art_col+str(cell)].value)
                # Remove twitter tags on the site
                soup = BeautifulSoup(store_site.text, 'html.parser')
                all_tags = soup.find_all('blockquote')
                for tag in all_tags:
                    if 'twitter' in str(tag):
                        tag.decompose()
                soup = str(soup)
                # Remove spaces and replace special characters with normal versions
                site_text = unicodedata.normalize('NFKD', soup)\
                            .encode('ASCII', 'ignore')\
                            .decode('UTF-8')
                site_text = cs(site_text)
                
            ### Test case:
                ### site_text = test_string
                ### print('\nstring being searched: ', site_text)
                
                # For each article, cycle through the saved tags for each brand
                for brand in HomePage.bran_list_1:
                    ## print('\nbrand = ', brand)
                    # Conditions dictionary key. Each row of tags is evaluated separately. Loop produces following dictionary: {'Project Runway SA': {1: False, 2: True, 3: False}, 'Edgars Fashion': ...}
                    tag_rows = 0
                    for position, value in zip(list(TagPage.saved_tags[brand].keys()), list(TagPage.saved_tags[brand].values())):
                        ## print('\nposition = ', position)
                        ## print('value = ', value)
                        # Postiion = dictionary key values, for example, '00' or '01'. Statement differentiates between entry widgets created by pressing and/or on TagPage. And in so doing, determines how the value in that entry should be looked for in the text of article.  
                        if position[1]!='0': # e.g., 01 and 02, or 21 and 22
                            # Add onto evaluation starting with or
                            cons_dict[brand][tag_rows] = cons_dict[brand][tag_rows] and cs(value) in site_text
                            ## print('if bloc:')
                            ## print('conditions dictionary = ', cons_dict)
                        else:
                            # Create new key in conditions dictionary that represents a row 
                            tag_rows += 1
                            cons_dict[brand][tag_rows] = cs(value) in site_text
                            ## print('else bloc triggered:')
                            ## print('conditions dictionary = ', cons_dict)

                    # If even one of the rows is evaluated true, then the ultimate evaluation for that brand will be true
                    sum_cons[brand] = cons_dict[brand][1]
                    ## print('cons_dict[{}][1] or row 1 evaluation: '.format(brand), cons_dict[brand][1])
                    ## print('initial value of sum_cons[{}]: '.format(brand), sum_cons[brand])
                    for row_num in range(2, tag_rows+1):
                        ## print('\ncons_dict key = ', row_num)
                        sum_cons[brand] = sum_cons[brand] or cons_dict[brand][row_num]
                        ## print('cons_dict[{}][{}] or row {} evaluation: '.format(brand, row_num, row_num), cons_dict[brand][row_num])
                        ## print('updated sum_cons[{}] value: '.format(brand), sum_cons[brand])
                
                for brand in HomePage.bran_list_1:
                    ## print('\nfinal sum_cons[{}] value: '.format(brand), sum_cons[brand])
                    ## print('brand tags being looked for = ', brand)
                    ## print('brand in cell = ', HomePage.ws1[brand_col+str(cell)].value)
                    # If tags found and brand being looked for is the brand in question
                    # Second condition added because you want a cell to be highlighted green based on the fact that project runway tags were being looked for 
                    # For-if loop needed because it ensures that only brands that had tags entered for them and thus appear in sum_cons are evaluated
                    # Without the second condition, sum_cons[brand] would return true even if it wasn't searching for that row's brand 
                    if (sum_cons[brand]) and (HomePage.ws1[brand_col+str(cell)].value == brand): 
                        # Note:
                        """
                        If next row has same site but different brand, am I having it do the search 
                        allover again? Answer appears to be yes. *** 
                        """
                        ## print('if bloc triggered')
                        # Highlight the article cell green
                        HomePage.ws1[art_col + str(cell)].fill = self.green_fill
                        break
                else:
                    ## print('else bloc triggered')
                    # Highlight the article cell red
                    HomePage.ws1[art_col + str(cell)].fill = self.red_fill
                
                ## print('\nSearch for unlisted brands, 1 = yes, 0 = no: {}'.format(SearchPage.v.get()))
                if SearchPage.v.get():
                    ## print('if bloc triggered')
                    ## print('Tag eval for each brand = ', sum_cons)
                    for condition, brand in zip(sum_cons.values(), HomePage.bran_list_1):
                        ## print('brand = {}, brand tags found: {}'.format(brand, condition))
                        
                        # condition: If brand tags for the brand in bran_list_1 were found
                        # (HomePage.ws1[brand_col+str(cell)].value != brand): And that brand was not the one being looked for
                        # (HomePage.ws1[art_col+str(cell)].value != HomePage.ws1[art_col+str(cell-1)].value): And the current article being searched does not equal the previous one. Prevents the same article being done twice.
                        # brand not in dfR_filt.loc[dfR_filt['Article URL']==HomePage.ws1[art_col+str(cell)].value,'Sub-Category Name'].tolist(): Sometimes article appears multiple times. Have to makes sure that brand won't be eventually searched for.  
                        if (condition) and (HomePage.ws1[brand_col+str(cell)].value != brand)\
                        and (HomePage.ws1[art_col+str(cell)].value != HomePage.ws1[art_col+str(cell-1)].value)\
                        and brand not in HomePage.dfR_filt.loc[HomePage.dfR_filt['Article URL']==HomePage.ws1[art_col+str(cell)].value,'Sub-Category Name'].tolist():

                            # Insert a row
                            HomePage.ws1.insert_rows(cell+1, amount=1)
                            # Put unlisted brand in brand column of new row
                            HomePage.ws1[brand_col+str(cell+1)] = brand
                            # Put article in previous row in article column of new row
                            HomePage.ws1[art_col+str(cell+1)] = HomePage.ws1[art_col+str(cell)].value
                            # Color article yellow to show it was added during the search
                            HomePage.ws1[brand_col+str(cell+1)].fill = self.yellow_fill
                else:
                    ## print('else bloc triggered')
                    pass

                # Wait half a second before looping. Prevents websites from detecting the bot.
                time.sleep(0.5)
            # Highlight any row article where various network errors are encountered as blue
            except Exception as e:
                ## print('\nexcept')
                ## print('error: ', e)
                HomePage.ws1[art_col + str(cell)].fill = self.blue_fill

            #Update progress bar
            # Determines how much progress bar moves
            int_var.set(delta*(num+1))
            self.progress_bar['variable'] = int_var
            self.progress_label['text'] = '{}%'.format(int_var.get())
            self.update()
        # Save edited worksheet to a new excel file
        HomePage.BrandWB.save('/Users/SplitShiftKing/Documents/Career/Projects/ProjectRunway/PR_Bot/WithClasses/ProRunOutputFile.xlsx')
        # Display how long it took the search to run
        end_time = time.monotonic()
        self.durationLabel['text'] = timedelta(seconds=end_time - start_time)
        ## print('Runtime = ', timedelta(seconds=end_time - start_time))


### test_string = 'asdf'    
### test range: 35-39

# Create instance of HomePage class
startPage = HomePage(InWin)
# Launch app
InWin.mainloop()
